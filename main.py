import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
import string

#Чтение файла Excel
excel_file=pd.read_excel("supermarket_sales - Sheet1.xlsx")
excel_file[["Gender","Product line","Total"]]

#Создание сводной таблицы
report_table = excel_file.pivot_table(index='Gender',
                                      columns='Product line',
                                      values='Total',
                                      aggfunc='sum').round(0)

#Экспорт сводной таблицы в файл Excel
report_table.to_excel('report_2021.xlsx',
                      sheet_name='Report',
                      startrow=4)

#Создание ссылки на строку и столбец
wb = load_workbook('report_2021.xlsx')
sheet = wb['Report']
#Cсылки на ячейки (исходная электронная таблица)
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

#Давайте выведем на экран созданные нами переменные, чтобы понять, что они означают. В данном случае мы получим следующие числа:
Min Columns: 1
Max Columns: 7
Min Rows: 5
Max Rows: 7

#Добавление диаграмм в Excel при помощи Python
wb = load_workbook('report_2021.xlsx')
sheet = wb['Report']

# barchart
barchart = BarChart()

#найти данные и категории
data = Reference(sheet,
                 min_col=min_column+1,
                 max_col=max_column,
                 min_row=min_row,
                 max_row=max_row) #including headers
categories = Reference(sheet,
                       min_col=min_column,
                       max_col=min_column,
                       min_row=min_row+1,
                       max_row=max_row) #not including headers
#добавление данных и категорий
barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

#Cхема местоположения
sheet.add_chart(barchart, "B12")
barchart.title = 'Sales by Product line'
barchart.style = 3

#выбрать стиль диаграммы
wb.save('report_2021.xlsx')

# Объяснение кода:
# barchart = BarChart() инициализирует переменную barchart из класса Barchart.
# data и categories – это переменные, которые показывают, где находится необходимая информация. Для автоматизации мы используем ссылки на столбцы и строки, которые определили выше. Также имейте в виду, что мы включаем заголовки в данные, но не в категории.
# Мы используем add_data() и set_categories(), чтобы добавить необходимые данные в гистограмму. Внутри add_data() добавим title_from_data = True, потому что мы включили заголовки для данных.
# Метод sheet.add_chart() используется для указания, что мы хотим добавить нашу гистограмму в лист Report. Также мы указываем, в какую ячейку мы хотим её добавить.
# Дальше мы изменяем заголовок и стиль диаграммы, используя barchart.title и barchart.style.
# И наконец, сохраняем все изменения с помощью wb.save()



# Применение формул в Excel через Python
# Предположим, мы хотим суммировать данные в ячейках B5 и B6 и отображать их в ячейке B7. 
# Кроме того, мы хотим установить формат ячейки B7 как денежный. Сделать мы это можем следующим образом:

sheet['B7'] = '=SUM(B5:B6)'
sheet['B7'].style = 'Currency'

# Однако сначала нам нужно получить алфавит, чтобы ссылаться на столбцы в Excel (A, B, C, …). Для этого воспользуемся библиотекой строк и напишем следующий код:
import string
alphabet = list(string.ascii_uppercase)
excel_alphabet = alphabet[0:max_column]
print(excel_alphabet)

#Применение формулы к нескольким ячейкам

sheet['B7'] = '=SUM(B5:B6)'
sheet['B7'].style = 'Currency'

print(min_column)
print(max_column)
print(min_row)
print(max_row)
#мы используем ссылки на столбцы и помещаем их в цикл for:
wb = load_workbook('report_2021.xlsx')
sheet = wb['Report']
# сумма в столбцах B-G
for i in excel_alphabet:
    if i!='A':
        sheet[f'{i}{max_row+1}'] = f'=SUM({i}{min_row+1}:{i}{max_row})'
        sheet[f'{i}{max_row+1}'].style = 'Currency'
# добавление общей метки
sheet[f'{excel_alphabet[0]}{max_row+1}'] = 'Total'
wb.save('report_2021.xlsx')

#for i in excel_alphabet проходит по всем активным столбцам, кроме столбца A (if i! = 'A'), так как столбец A не содержит числовых данных
#запись sheet[f'{i}{max_row+1}'] = f'=SUM({i}{min_row+1}:{i}{max_row}'  это то же самое, что и sheet['B7'] = '=SUM(B5:B6)', только для столбцов от A до G
#строчка sheet [f '{i} {max_row + 1}'].style = 'Currency' задает денежный формат ячейкам с числовыми данными (т.е. тут мы опять же исключаем столбец А)
#мы добавляем запись Total в столбец А под максимальной строкой (т.е. под седьмой), используя код [f '{excel_alphabet [0]} {max_row + 1}'] = 'Total'

#Форматирование листа с отчетом
wb = load_workbook('report_2021.xlsx')
sheet = wb['Report']
sheet['A1'] = 'Sales Report'
sheet['A2'] = '2021'
sheet['A1'].font = Font('Arial', bold=True, size=20)
sheet['A2'].font = Font('Arial', bold=True, size=10)
wb.save('report_2021.xlsx')

