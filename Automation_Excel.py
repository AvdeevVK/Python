#Автоматизация отчета с помощью функции Python
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
import string

def automate_excel(file_name):
    """The file name should have the following structure: sales_month.xlsx"""
    # прочитать эксель файл
    excel_file = pd.read_excel(file_name)
    # сделать сводную таблицу
    report_table = excel_file.pivot_table(index='Gender', columns='Product line', values='Total', aggfunc='sum').round(0)
    # отделение месяца и расширения от имени файла
    month_and_extension = file_name.split('_')[1]
    # отправить таблицу отчета в файл excel
    report_table.to_excel(f'report_{month_and_extension}', sheet_name='Report', startrow=4)
    # загрузка книги и выбор листа
    wb = load_workbook(f'report_{month_and_extension}')
    sheet = wb['Report']
    # ссылки на ячейки (исходная электронная таблица)
    min_column = wb.active.min_column
    max_column = wb.active.max_column
    min_row = wb.active.min_row
    max_row = wb.active.max_row
    # добавление диаграммы
    barchart = BarChart()
    data = Reference(sheet, min_col=min_column+1, max_col=max_column, min_row=min_row, max_row=max_row) #including headers
    categories = Reference(sheet, min_col=min_column, max_col=min_column, min_row=min_row+1, max_row=max_row) #not including headers
    barchart.add_data(data, titles_from_data=True)
    barchart.set_categories(categories)
    sheet.add_chart(barchart, "B12") #location chart
    barchart.title = 'Sales by Product line'
    barchart.style = 2 #choose the chart style
    # применение формул
    # сначала создайте список алфавитов в качестве ссылок для ячеек
    alphabet = list(string.ascii_uppercase)
    excel_alphabet = alphabet[0:max_column] #note: Python lists start on 0 -> A=0, B=1, C=2. #note2 the [a:b] takes b-a elements
    # сумма в столбцах B-G
    for i in excel_alphabet:
        if i!='A':
            sheet[f'{i}{max_row+1}'] = f'=SUM({i}{min_row+1}:{i}{max_row})'
            sheet[f'{i}{max_row+1}'].style = 'Currency'
    sheet[f'{excel_alphabet[0]}{max_row+1}'] = 'Total'
    # получение названия месяца
    month_name = month_and_extension.split('.')[0]
    # форматирование отчета
    sheet['A1'] = 'Sales Report'
    sheet['A2'] = month_name.title()
    sheet['A1'].font = Font('Arial', bold=True, size=20)
    sheet['A2'].font = Font('Arial', bold=True, size=10)
    wb.save(f'report_{month_and_extension}')
    return print('report done!')

automate_excel('sales_2021.xlsx')
  
  
